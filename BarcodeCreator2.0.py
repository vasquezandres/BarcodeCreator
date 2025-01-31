import os
import tkinter as tk
from tkinter import messagebox, filedialog
import barcode
from barcode.writer import ImageWriter
from PIL import Image, ImageTk, ImageDraw, ImageFont
import random
import csv
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import subprocess
import webbrowser

# Variable global para la carpeta de destino
selected_folder = ""
# Lista para almacenar los códigos generados y sus detalles
generated_codes = []

# Rutas a los archivos de ícono y logo
icon_path = "logo.ico"  # Cambia esto por la ruta correcta
logo_path = "logo.png"  # Cambia esto por la ruta correcta


def select_folder():
    global selected_folder
    folder = filedialog.askdirectory()
    if folder:
        selected_folder = folder
        folder_label.config(text=f"Carpeta seleccionada:\n{folder}")
        load_existing_codes()


def regenerate_existing_barcodes():
    if not selected_folder:
        messagebox.showwarning("Advertencia", "Por favor, selecciona una carpeta primero.")
        return
    for item in generated_codes:
        generate_barcode_from_existing(item['code'], item['product_name'])


def generate_barcode_from_existing(code, product_name):
    try:
        filepath = os.path.join(selected_folder, f"barcode_{code}.png")
        code128 = barcode.get_barcode_class('code128')
        barcode_instance = code128(code, writer=ImageWriter())
        barcode_instance.save(filepath)

        img = Image.open(filepath)
        resize_option = (150, 300) if resize_small.get() else (272, 280)
        img = img.resize(resize_option, Image.LANCZOS)

        if include_product_name.get():
            draw = ImageDraw.Draw(img)
            font = ImageFont.load_default()
            text_position = (10, img.height - 20)
            draw.text(text_position, product_name, fill="black", font=font)
        img.save(filepath)
    except Exception as e:
        messagebox.showerror("Error", f"Error al regenerar código {code}:\n{e}")


def load_existing_codes():
    global generated_codes
    csv_path = os.path.join(selected_folder, "codigos_generados.csv")
    generated_codes.clear()
    if os.path.exists(csv_path):
        with open(csv_path, newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            next(reader, None)
            for row in reader:
                if len(row) >= 2:
                    generated_codes.append({"code": row[0], "product_name": row[1], "filepath": row[2]})


def generate_random_code():
    prefix = "USA" if use_prefix.get() else ""
    random_code = prefix + ''.join(random.choices('0123456789', k=10 if not use_prefix.get() else 6))
    entry_code.delete(0, tk.END)
    entry_code.insert(0, random_code)


def generate_barcode():
    include_name = include_product_name.get()
    global selected_folder
    if not selected_folder:
        messagebox.showwarning("Advertencia", "Por favor, selecciona una carpeta primero.")
        return

    code = entry_code.get()
    product_name = entry_name.get()

    if not code:
        generate_random_code()
        code = entry_code.get()

    if not product_name:
        messagebox.showwarning("Advertencia", "Por favor, ingresa un nombre de producto.")
        return

    if any(item["code"] == code for item in generated_codes):
        messagebox.showwarning("Advertencia", "El código ya existe. Usa otro o genera uno nuevo.")
        return

    try:
        os.makedirs(selected_folder, exist_ok=True)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo crear la carpeta:\n{selected_folder}\nError: {e}")
        return

    filepath = os.path.join(selected_folder, f"barcode_{code}.png")
    temp_filepath = os.path.join(selected_folder, f"temp_barcode_{code}.png")

    try:
        code128 = barcode.get_barcode_class('code128')
        barcode_instance = code128(code, writer=ImageWriter())
        temp_filepath = barcode_instance.save("temp_barcode")
        os.rename(temp_filepath, filepath)

        if not os.path.exists(filepath):
            raise FileNotFoundError(f"No se encontró el archivo generado: {filepath}")

        img = Image.open(filepath)
        resize_option = (150, 300) if resize_small.get() else (272, 280)
        img = img.resize(resize_option, Image.LANCZOS)

        if include_name:
            draw = ImageDraw.Draw(img)
            font = ImageFont.load_default()
            text_position = (10, img.height - 20)
            draw.text(text_position, product_name, fill="black", font=font)
        img.save(filepath)  # Guarda la imagen con el nuevo tamaño

        img_tk = ImageTk.PhotoImage(img)
        barcode_label.config(image=img_tk)
        barcode_label.image = img_tk

        generated_codes.append({"code": code, "product_name": product_name, "filepath": filepath})
        save_to_csv()

        messagebox.showinfo("Éxito", f"Código de barras generado y guardado en:\n{filepath}")
    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al generar el código de barras:\n{e}")


def save_to_csv():
    csv_path = os.path.join(selected_folder, "codigos_generados.csv")
    with open(csv_path, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(["Código", "Nombre del Producto", "Ruta de la Imagen"])
        for item in generated_codes:
            writer.writerow([item["code"], item["product_name"], item["filepath"]])


def print_barcode(filepath):
    try:
        subprocess.run(["mspaint", filepath], check=True)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo imprimir el código de barras:\n{e}")


def export_all_barcodes():
    if not selected_folder:
        messagebox.showwarning("Advertencia", "Por favor, selecciona una carpeta primero.")
        return

    filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if filepath:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Códigos de Barras"
            ws.append(["Código", "Nombre del Producto", "Imagen del Código de Barras"])

            # Buscar todas las imágenes de códigos de barras en la carpeta seleccionada
            for filename in os.listdir(selected_folder):
                if filename.endswith(".png") and "barcode_" in filename:
                    code = filename.replace("barcode_", "").replace(".png", "")
                    product_name = next((item['product_name'] for item in generated_codes if item['code'] == code), "Desconocido")
                    ws.append([code, product_name, ""])
                    img = ExcelImage(os.path.join(selected_folder, filename))
                    img.width = 200
                    img.height = 60
                    ws.add_image(img, f"C{ws.max_row}")

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 25

            wb.save(filepath)
            messagebox.showinfo("Éxito", f"Todos los códigos de barras han sido exportados a {filepath}")
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al exportar a Excel: {e}")


def export_to_excel():
    if not generated_codes:
        messagebox.showwarning("Advertencia", "No hay códigos de barras generados para exportar.")
        return

    filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if filepath:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Códigos de Barras"
            ws.append(["Código", "Nombre del Producto", "Imagen del Código de Barras"])

            for item in generated_codes:
                ws.append([item["code"], item["product_name"], ""])
                img = ExcelImage(item["filepath"])
                img.width = 200
                img.height = 60
                ws.add_image(img, f"C{ws.max_row}")

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 25

            wb.save(filepath)
            messagebox.showinfo("Éxito", f"Lista exportada a {filepath}")
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al exportar a Excel: {e}")


# Interfaz gráfica
root = tk.Tk()
root.title("Generador de Códigos de Barras")

# Ajustar el tamaño de la ventana
root.geometry("800x600")  # Ancho x Alto

# Establecer el ícono de la ventana
if os.path.exists(icon_path):
    root.iconbitmap(icon_path)
else:
    messagebox.showwarning("Advertencia", "No se encontró el archivo de ícono.")

# Cargar y mostrar el logo
if os.path.exists(logo_path):
    logo_image = Image.open(logo_path)
    logo_image = logo_image.resize((100, 100), Image.LANCZOS)  # Ajusta el tamaño si es necesario
    logo_tk = ImageTk.PhotoImage(logo_image)
    logo_label = tk.Label(root, image=logo_tk)
    logo_label.image = logo_tk  # Evitar que la imagen sea eliminada por el recolector de basura
    logo_label.pack(pady=10)
else:
    messagebox.showwarning("Advertencia", "No se encontró el archivo de logo.")

# Variable para definir si se usa el prefijo "USR"
use_prefix = tk.BooleanVar(value=False)

folder_button = tk.Button(root, text="Seleccionar Carpeta", command=select_folder)
folder_button.pack(pady=10)

folder_label = tk.Label(root, text="No se ha seleccionado ninguna carpeta", wraplength=400)
folder_label.pack(pady=5)

use_prefix_check = tk.Checkbutton(root, text="Usar prefijo 'USA'", variable=use_prefix)
use_prefix_check.pack(pady=5)

label_code = tk.Label(root, text="Ingresa el código:")
label_code.pack(pady=5)

entry_code = tk.Entry(root)
entry_code.pack(pady=5)

label_name = tk.Label(root, text="Ingresa el nombre del producto:")
label_name.pack(pady=5)

entry_name = tk.Entry(root)
entry_name.pack(pady=5)

generate_button = tk.Button(root, text="Generar Código de Barras", command=generate_barcode)
generate_button.pack(pady=10)

random_button = tk.Button(root, text="Generar Código Aleatorio", command=generate_random_code)
random_button.pack(pady=10)

export_excel_button = tk.Button(root, text="Exportar a Excel", command=export_to_excel)
export_excel_button.pack(pady=10)

export_all_button = tk.Button(root, text="Exportar Todos los Códigos", command=export_all_barcodes)
export_all_button.pack(pady=10)

barcode_label = tk.Label(root)
resize_small = tk.BooleanVar(value=False)
resize_check = tk.Checkbutton(root, text="Usar tamaño 150x300", variable=resize_small)
resize_check.pack(pady=5)
include_product_name = tk.BooleanVar(value=False)
include_name_check = tk.Checkbutton(root, text="Incluir nombre en código", variable=include_product_name)
include_name_check.pack(pady=5)
regenerate_button = tk.Button(root, text="Regenerar Códigos Existentes", command=regenerate_existing_barcodes)
regenerate_button.pack(pady=10)
barcode_label.pack(pady=20)

# Agregar el enlace en la parte inferior
footer_label = tk.Label(
    root,
    text="Solutech Panamá | www.solutechpanama.com",
    fg="blue",
    cursor="hand2"  # Cambia el cursor a una mano para simular un enlace
)
footer_label.pack(side=tk.BOTTOM, pady=10)

# Opcional: Agregar un evento para abrir el enlace en el navegador
def open_website(event):
    webbrowser.open("http://www.solutechpanama.com")

footer_label.bind("<Button-1>", open_website)

root.mainloop()
